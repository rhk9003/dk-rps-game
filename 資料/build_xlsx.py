import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('final_data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

names = {
    'A0220': 'DK 動態釋壓鞋',
    'A0119': '鋅離子抗菌除臭五趾隱形襪',
    'A0003': '石墨烯能量塑形袖套',
    'A0114': '涼感底紗襪',
}

prices = {
    'A0220': 1280,
    'A0119': 250,
    'A0003': 680,
    'A0114': 180,
}

base_order = ['A0220', 'A0119', 'A0003', 'A0114']

def aggregate(base):
    qty = 0
    val = 0
    fz = 0
    color_count = 0
    for code, d in data.items():
        if code.startswith(base):
            color_count += 1
            if d['total']:
                qty += d['total'][0]
                val += d['total'][1]
            else:
                qty += sum(b[2] for b in d['branches'])
                val += sum(b[3] for b in d['branches'])
            fuzhong = [b for b in d['branches'] if b[0] == '002']
            if fuzhong:
                fz += fuzhong[0][2]
    return int(qty), int(val), int(fz), color_count

wb = Workbook()
thin = Side(border_style='thin', color='999999')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

win_fill = PatternFill('solid', fgColor='FFF2CC')
lose_fill = PatternFill('solid', fgColor='DEEBF7')
win_banner = PatternFill('solid', fgColor='BF8F00')
lose_banner = PatternFill('solid', fgColor='2E75B6')

# ==============================================================
# Sheet 1: 獎項總表
# ==============================================================
ws = wb.active
ws.title = '獎項總表'

# --- 標題 ---
ws['A1'] = 'DK 府中店滿月｜閨密機猜拳遊戲 獎項總表'
ws['A1'].font = Font(bold=True, size=16)
ws.merge_cells('A1:L1')

ws['A2'] = ('設計基準：30 天 × 25 人/日 = 750 場 ／ 贏率 50% ／ 券走 LINE bot 發送不限張數'
            ' ／ 資料來源：庫存結構透視分析-20260417.xls、府中店櫃銷售多維分析-20260416.xls')
ws['A2'].font = Font(size=10, color='666666')
ws['A2'].alignment = Alignment(wrap_text=True)
ws.merge_cells('A2:L2')
ws.row_dimensions[2].height = 32

# =============================================================
# 區塊 1（放最前面）：📌 機率與數量分配（討論用精簡表）
# =============================================================
row = 4
c = ws.cell(row=row, column=1, value='📌 機率與數量分配（討論用主表）')
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='C00000')
c.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws.row_dimensions[row].height = 24
row += 1

headers = ['獎項分類', '類別', '編號/券種', '項目名稱', '規格/說明',
           '單價/面額', '機率', '數量上限', '每日上限', '全公司庫存', '府中店庫存', '備註']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
row += 1

# ---- 贏家 ----
c = ws.cell(row=row, column=1, value='🏆 贏家獎（贏家池 100%）')
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = win_banner
c.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws.row_dimensions[row].height = 22
row += 1

win_products = [
    ('A0220', 8.0, 30, '每日 1 雙', '熱賣款；府中 98 雙'),
    ('A0119', 25.0, 60, '每日 2 雙', '熱銷+庫存充裕；府中 91 雙'),
]
for base, prob, cap, daily_cap, note in win_products:
    qty, val, fz, cc_count = aggregate(base)
    spec = f'共 {cc_count} 色（先不分尺寸）'
    vals = ['贏家獎', '實體商品', base, names[base], spec,
            f'${prices[base]:,}', f'{prob}%', f'{cap} 件', daily_cap,
            f'{qty:,} 件', f'{fz} 件', note]
    for i, v in enumerate(vals, 1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.fill = win_fill
        cc.border = border
        cc.alignment = Alignment(horizontal='center' if i in [1, 2, 3, 6, 7, 8, 9, 10, 11] else 'left',
                                 vertical='center', wrap_text=True)
    row += 1

win_coupons = [
    ('LINE $500 官網抵用金', '電商券', '官網買原價鞋款(不含拖鞋)',
     '$500', 20.0, '滿 $1,800 折抵；~5/31'),
    ('LINE 門市 7 折券', '門市券', '限原價商品', '7 折', 25.0, '~5/31；不得與其他優惠併用'),
    ('LINE 門市 $500 折價券', '門市券', '限原價商品', '$500', 22.0, '~5/31；不得與其他優惠併用'),
]
for name_, ctype, scope, face, prob, note in win_coupons:
    vals = ['贏家獎', '優惠券', ctype, name_, scope, face, f'{prob}%', '不限', '不限',
            'LINE bot', 'LINE bot', note]
    for i, v in enumerate(vals, 1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.fill = win_fill
        cc.border = border
        cc.alignment = Alignment(horizontal='center' if i in [1, 2, 6, 7, 8, 9, 10, 11] else 'left',
                                 vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 28
    row += 1

# 贏家小計
vals = ['贏家獎', '小計', '', '', '', '', '100%', '90 件實體+不限券', '', '', '', '贏家池總機率']
for i, v in enumerate(vals, 1):
    cc = ws.cell(row=row, column=i, value=v)
    cc.font = Font(bold=True)
    cc.fill = PatternFill('solid', fgColor='F4B084')
    cc.border = border
    cc.alignment = Alignment(horizontal='center' if i in [1, 2, 7, 8] else 'left', vertical='center')
row += 1

# ---- 輸家 ----
c = ws.cell(row=row, column=1, value='🎁 輸家獎／參加獎（輸家池 100%）')
c.font = Font(bold=True, size=12, color='FFFFFF')
c.fill = lose_banner
c.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws.row_dimensions[row].height = 22
row += 1

lose_products = [
    ('A0003', 2.5, 10, '抽完為止', '🎊 稀有隱藏驚喜！全公司僅 16 件、府中 0（需調 10 件）；抽完自動改發 A0114'),
    ('A0114', 20.0, 60, '每日 2 雙', '府中 22 雙，不足再從他店調'),
]
for base, prob, cap, daily_cap, note in lose_products:
    qty, val, fz, cc_count = aggregate(base)
    spec = f'共 {cc_count} 色（先不分尺寸）'
    vals = ['輸家獎', '實體商品', base, names[base], spec,
            f'${prices[base]:,}', f'{prob}%', f'{cap} 件', daily_cap,
            f'{qty:,} 件', f'{fz} 件', note]
    for i, v in enumerate(vals, 1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.fill = lose_fill
        cc.border = border
        cc.alignment = Alignment(horizontal='center' if i in [1, 2, 3, 6, 7, 8, 9, 10, 11] else 'left',
                                 vertical='center', wrap_text=True)
    row += 1

lose_coupons = [
    ('LINE $50 官網抵用金', '電商券', '官網買鞋款/包類', '$50', 20.0, '滿 $51 折抵；~5/31'),
    ('LINE 門市 8 折券', '門市券', '限原價商品', '8 折', 27.5, '~5/31；不得與其他優惠併用'),
    ('LINE 門市 $100 折價券', '門市券', '限原價商品', '$100', 30.0, '~5/31；不得與其他優惠併用'),
]
for name_, ctype, scope, face, prob, note in lose_coupons:
    vals = ['輸家獎', '優惠券', ctype, name_, scope, face, f'{prob}%', '不限', '不限',
            'LINE bot', 'LINE bot', note]
    for i, v in enumerate(vals, 1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.fill = lose_fill
        cc.border = border
        cc.alignment = Alignment(horizontal='center' if i in [1, 2, 6, 7, 8, 9, 10, 11] else 'left',
                                 vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 28
    row += 1

vals = ['輸家獎', '小計', '', '', '', '', '100%', '70 件實體+不限券', '', '', '', '輸家池總機率']
for i, v in enumerate(vals, 1):
    cc = ws.cell(row=row, column=i, value=v)
    cc.font = Font(bold=True)
    cc.fill = PatternFill('solid', fgColor='9DC3E6')
    cc.border = border
    cc.alignment = Alignment(horizontal='center' if i in [1, 2, 7, 8] else 'left', vertical='center')
row += 2

# =============================================================
# 區塊 2：💰 成本試算
# =============================================================
c = ws.cell(row=row, column=1, value='💰 成本試算（750 場 × 贏率 50%）')
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='548235')
c.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws.row_dimensions[row].height = 24
row += 1

cost_headers = ['分類', '項目', '單價', '數量上限', '最大成本', '期望實發', '期望成本', '備註', '', '', '', '']
for i, h in enumerate(cost_headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='70AD47')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
row += 1

pool_size = 375

cost_items = [
    ('贏家獎', 'A0220 動態釋壓鞋', 1280, 30, 8.0, win_fill),
    ('贏家獎', 'A0119 五趾隱形襪', 250, 60, 25.0, win_fill),
    ('輸家獎', 'A0003 石墨烯袖套', 680, 10, 2.5, lose_fill),
    ('輸家獎', 'A0114 涼感底紗襪', 180, 60, 20.0, lose_fill),
]

total_max_cost = 0
total_exp_cost = 0
for cat, item, unit, cap, prob, fill in cost_items:
    max_cost = unit * cap
    exp_qty = pool_size * prob / 100
    exp_qty_capped = min(exp_qty, cap)
    exp_cost = int(unit * exp_qty_capped)
    total_max_cost += max_cost
    total_exp_cost += exp_cost
    note = f'期望 {exp_qty:.1f} 件' + ('（觸頂）' if exp_qty > cap else '')
    vals = [cat, item, f'${unit:,}', f'{cap}', f'${max_cost:,}', f'{exp_qty_capped:.1f}',
            f'${exp_cost:,}', note, '', '', '', '']
    for i, v in enumerate(vals, 1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.fill = fill
        cc.border = border
        cc.alignment = Alignment(horizontal='center' if i in [1, 3, 4, 5, 6, 7] else 'left', vertical='center')
    row += 1

vals = ['合計', '實體獎品', '—', '160 件', f'${total_max_cost:,}', '—',
        f'${total_exp_cost:,}', 'LINE 券邊際成本 ≈ 0', '', '', '', '']
for i, v in enumerate(vals, 1):
    cc = ws.cell(row=row, column=i, value=v)
    cc.font = Font(bold=True)
    cc.fill = PatternFill('solid', fgColor='C6E0B4')
    cc.border = border
    cc.alignment = Alignment(horizontal='center' if i in [1, 3, 4, 5, 6, 7] else 'left', vertical='center')
row += 1

vals = ['', '每場平均獎品成本', '', '', f'${total_max_cost/750:.0f}/場（上限）', '',
        f'${total_exp_cost/750:.0f}/場（期望）', '', '', '', '', '']
for i, v in enumerate(vals, 1):
    cc = ws.cell(row=row, column=i, value=v)
    cc.font = Font(italic=True, color='666666')
    cc.border = border
    cc.alignment = Alignment(horizontal='center' if i in [5, 7] else 'left', vertical='center')
row += 2

# =============================================================
# 區塊 3：📊 場次敏感度
# =============================================================
c = ws.cell(row=row, column=1, value='📊 場次敏感度（若實際場次不同）')
c.font = Font(bold=True, size=13, color='FFFFFF')
c.fill = PatternFill('solid', fgColor='7030A0')
c.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
ws.row_dimensions[row].height = 24
row += 1

sens_hdrs = ['情境', '日均玩家', '總場次', '贏家場次', 'A0220 期望中獎', '實體成本期望', '備註', '', '', '', '', '']
for i, h in enumerate(sens_hdrs, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='7030A0')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
row += 1

scenarios = [
    ('保守', 20, 600, '業績淡季+陰雨天'),
    ('基準', 25, 750, '依 3/28~4/16 平均'),
    ('樂觀', 35, 1050, '週末+媒體曝光帶動'),
]
for label, per_day, total_game, note in scenarios:
    winners = total_game * 0.5
    a0220_exp = min(winners * 0.08, 30)
    est = 0
    w_pool = total_game * 0.5
    l_pool = total_game * 0.5
    for cat, _, unit, cap, prob, _ in cost_items:
        pool = w_pool if cat == '贏家獎' else l_pool
        exp_qty = min(pool * prob / 100, cap)
        est += unit * exp_qty
    vals = [label, f'{per_day} 人', f'{total_game} 場', f'{int(winners)} 場',
            f'{int(a0220_exp)} 雙', f'${int(est):,}', note, '', '', '', '', '']
    fill = PatternFill('solid', fgColor='E4DFEC' if label == '基準' else 'F2F2F2')
    for i, v in enumerate(vals, 1):
        cc = ws.cell(row=row, column=i, value=v)
        cc.fill = fill
        cc.border = border
        if label == '基準':
            cc.font = Font(bold=True)
        cc.alignment = Alignment(horizontal='center' if i in [1, 2, 3, 4, 5, 6] else 'left', vertical='center')
    row += 1

widths = [12, 12, 18, 24, 24, 12, 10, 14, 14, 14, 12, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A5'


# ==============================================================
# Sheet 2: 府中店業績參考
# ==============================================================
ws_f = wb.create_sheet('府中店業績參考')
ws_f['A1'] = '府中店重開期業績與遊戲場次估算依據'
ws_f['A1'].font = Font(bold=True, size=14)
ws_f.merge_cells('A1:F1')

ws_f['A2'] = '資料來源：府中店櫃銷售多維分析-20260416.xls（3/23 重開 ~ 4/16）'
ws_f['A2'].font = Font(size=10, color='666666')
ws_f.merge_cells('A2:F2')

import xlrd
from collections import defaultdict
from datetime import datetime
weekdays = ['一', '二', '三', '四', '五', '六', '日']

daily_agg = {}
try:
    wb_f = xlrd.open_workbook('../../府中分析/府中店櫃銷售多維分析-20260416.xls', encoding_override='cp950')
    ws_src = wb_f.sheet_by_index(0)
    daily_agg_d = defaultdict(lambda: [0, 0, 0])
    for i in range(4, ws_src.nrows):
        date_val = ws_src.cell_value(i, 3)
        if not date_val:
            continue
        date_str = str(date_val).strip().split('.')[0]
        if len(date_str) == 8 and date_str.isdigit():
            qty = ws_src.cell_value(i, 4) or 0
            cust = ws_src.cell_value(i, 5) or 0
            amt = ws_src.cell_value(i, 6) or 0
            daily_agg_d[date_str][0] += qty
            daily_agg_d[date_str][1] += cust
            daily_agg_d[date_str][2] += amt
    daily_agg = daily_agg_d
except Exception as e:
    print('sales data read failed:', e)

recent = sorted([d for d in daily_agg.keys() if d >= '20260323'])

hdrs = ['日期', '星期', '銷售數量', '會員來客數', '營業額', '備註']
for i, h in enumerate(hdrs, 1):
    c = ws_f.cell(row=4, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center')
    c.border = border

r = 5
total_cust = 0
count = 0
for d in recent:
    row_data = daily_agg[d]
    date_obj = datetime.strptime(d, '%Y%m%d')
    wd = weekdays[date_obj.weekday()]
    is_weekend = date_obj.weekday() >= 5
    note = ''
    if d <= '20260327':
        note = '重開初期'
    elif is_weekend:
        note = '週末'
    fill_row = PatternFill('solid', fgColor='FFF2CC' if is_weekend else 'FFFFFF')
    vals = [f'{d[:4]}/{d[4:6]}/{d[6:]}', wd,
            int(row_data[0]), int(row_data[1]), f'${int(row_data[2]):,}', note]
    for i, v in enumerate(vals, 1):
        cc = ws_f.cell(row=r, column=i, value=v)
        cc.border = border
        cc.fill = fill_row
        cc.alignment = Alignment(horizontal='center' if i in [1, 2, 3, 4] else 'left', vertical='center')
    if d >= '20260328':
        total_cust += row_data[1]
        count += 1
    r += 1

r += 1
avg_cust = total_cust / count if count else 0
summary = [
    ('3/28 ~ 4/16 穩定期平均來客', f'{avg_cust:.1f} 人/日（有消費會員）'),
    ('預估實際進店（含散客）', f'{avg_cust * 1.5:.0f} ~ {avg_cust * 2:.0f} 人/日'),
    ('閨密機玩家估計', '25 人/日（假設 50% 進店客會玩）'),
    ('30 天活動總場次基準', '750 場'),
]
for label, val in summary:
    ws_f.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws_f.cell(row=r, column=2, value=val)
    ws_f.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1

for i, w in enumerate([14, 8, 12, 14, 16, 20], 1):
    ws_f.column_dimensions[get_column_letter(i)].width = w


# ==============================================================
# Sheet 3: 色號明細
# ==============================================================
color_map = {
    '04': '色04', '07': '色07', '20': '白', '27': '色27',
    '30': '綠', '32': '色32', '40': '黃', '50': '紫',
    '64': '色64', '67': '粉', '69': '灰', '70': '棕',
    '73': '藍', '90': '黑', '99': '綜合/多色',
}
category_map = {
    'A0220': '贏家獎', 'A0119': '贏家獎',
    'A0003': '輸家獎', 'A0114': '輸家獎',
}

ws2 = wb.create_sheet('色號明細（參考）')
ws2['A1'] = '各商品色號庫存（參考用）'
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:G1')

hdrs = ['分類', '商品編號', '商品名稱', '色號', '總庫存', '單價', '府中店庫存']
for i, h in enumerate(hdrs, 1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='4472C4')
    c.alignment = Alignment(horizontal='center')
    c.border = border

r = 4
for base in base_order:
    codes = sorted([c for c in data.keys() if c.startswith(base)])
    cat = category_map[base]
    fill = win_fill if '贏家' in cat else lose_fill
    for code in codes:
        suffix = code.split('-')[1]
        color = color_map.get(suffix, suffix)
        total = data[code]['total']
        qty = int(total[0]) if total else int(sum(b[2] for b in data[code]['branches']))
        fuzhong = [b for b in data[code]['branches'] if b[0] == '002']
        fz_qty = int(fuzhong[0][2]) if fuzhong else 0
        vals = [cat, code, names[base], color, qty, prices[base], fz_qty]
        for i, v in enumerate(vals, 1):
            cc = ws2.cell(row=r, column=i, value=v)
            cc.fill = fill
            cc.border = border
            cc.alignment = Alignment(horizontal='center' if i in [1, 2, 4, 5, 6, 7] else 'left')
        r += 1

for i, w in enumerate([12, 14, 26, 10, 10, 10, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ==============================================================
# Sheets 4+: 分店明細
# ==============================================================
for sheet_name, base in [('A0220_分店', 'A0220'), ('A0119_分店', 'A0119'),
                         ('A0003_分店', 'A0003'), ('A0114_分店', 'A0114')]:
    wss = wb.create_sheet(sheet_name)
    wss['A1'] = f'{base} {names[base]}｜各分店庫存明細'
    wss['A1'].font = Font(bold=True, size=14)
    wss.merge_cells('A1:F1')

    wss['A2'] = f'單價 $ {prices[base]} ／ 分類：{category_map[base]}'
    wss['A2'].font = Font(size=10, color='666666')
    wss.merge_cells('A2:F2')

    codes = sorted([c for c in data.keys() if c.startswith(base)])

    hdrs = ['商品編號', '色號', '分店編號', '分店名稱', '數量', '金額']
    for i, h in enumerate(hdrs, 1):
        c = wss.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='4472C4')
        c.alignment = Alignment(horizontal='center')
        c.border = border

    r = 5
    for code in codes:
        suffix = code.split('-')[1]
        color = color_map.get(suffix, suffix)
        branches = sorted(data[code]['branches'], key=lambda x: -x[2])
        for b in branches:
            vals = [code, color, b[0], b[1], int(b[2]), int(b[3])]
            for i, v in enumerate(vals, 1):
                cc = wss.cell(row=r, column=i, value=v)
                cc.border = border
                cc.alignment = Alignment(horizontal='center' if i in [1, 2, 3, 5, 6] else 'left')
                if b[0] == '002':
                    cc.fill = PatternFill('solid', fgColor='FFE699')
            r += 1
        total = data[code]['total']
        if total:
            vals = [code, color, '', '小計', int(total[0]), int(total[1])]
            for i, v in enumerate(vals, 1):
                cc = wss.cell(row=r, column=i, value=v)
                cc.font = Font(bold=True)
                cc.fill = PatternFill('solid', fgColor='D9E1F2')
                cc.border = border
                cc.alignment = Alignment(horizontal='center' if i in [1, 2, 3, 5, 6] else 'left')
            r += 1
        r += 1

    for i, w in enumerate([14, 10, 12, 20, 10, 12], 1):
        wss.column_dimensions[get_column_letter(i)].width = w

out_path = '../府中店滿月_贈品整理_20260417.xlsx'
wb.save(out_path)
import os
print('saved to:', os.path.abspath(out_path))
