import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/green_results.json', 'r', encoding='utf-8') as f:
    results = json.load(f)

wb_inv = openpyxl.load_workbook('資料/庫存明細20260415.xlsx', data_only=True)
ws_inv = wb_inv.active

inv = {}
for r in range(7, ws_inv.max_row + 1):
    code = ws_inv.cell(row=r, column=1).value
    if not code:
        continue
    code = str(code).strip()
    inv.setdefault(code, []).append({
        'name': str(ws_inv.cell(row=r, column=2).value or '').strip(),
        'site': str(ws_inv.cell(row=r, column=3).value or '').strip(),
        'price': ws_inv.cell(row=r, column=4).value,
        'qty': ws_inv.cell(row=r, column=6).value or 0,
        'amt': ws_inv.cell(row=r, column=7).value or 0,
    })

out_wb = openpyxl.Workbook()
ws1 = out_wb.active
ws1.title = '綠色不補號彙總'

title_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
cat_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
subtotal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
zero_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

ws1.merge_cells('A1:H1')
t = ws1.cell(row=1, column=1, value='門市周邊綠色不補號商品 — 價格與庫存彙總  (庫存明細製表日期 2026-04-15)')
t.font = Font(bold=True, color='FFFFFF', size=14)
t.fill = title_fill
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

headers = ['分類', '商品型號', '品名規格', '單價', '總庫存', '總庫存金額', '庫點數', '備註']
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col_idx, value=h)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws1.row_dimensions[2].height = 24

cat_order = ['襪子', '拖鞋', '射出拖鞋', '包包', '香氛', '鞋保養品', '鞋墊', '護具']
row_idx = 3
grand_qty = 0
grand_amt = 0

for cat in cat_order:
    cat_items = [r for r in results if r['cat'] == cat]
    cat_qty = 0
    cat_amt = 0
    for r in cat_items:
        name = r['inv_name'] if r['inv_name'] else r['pdf_name']
        name = name.replace('_x000D_', '').strip()
        price = r['price']
        qty = r['total_qty']
        amt = r['total_amt']
        sites = r['n_sites']
        if qty > 0:
            note = ''
        else:
            note = '庫存=' + str(qty) + ' (可能退貨/盤虧)'

        cells = [
            ws1.cell(row=row_idx, column=1, value=cat),
            ws1.cell(row=row_idx, column=2, value=r['code']),
            ws1.cell(row=row_idx, column=3, value=name),
            ws1.cell(row=row_idx, column=4, value=price),
            ws1.cell(row=row_idx, column=5, value=qty),
            ws1.cell(row=row_idx, column=6, value=amt),
            ws1.cell(row=row_idx, column=7, value=sites),
            ws1.cell(row=row_idx, column=8, value=note),
        ]
        for c in cells:
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cells[0].fill = cat_fill
        cells[0].font = Font(bold=True)
        cells[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[3].number_format = '#,##0'
        cells[4].number_format = '#,##0'
        cells[5].number_format = '#,##0'
        cells[7].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if qty <= 0:
            cells[4].fill = zero_fill
        cat_qty += qty
        cat_amt += amt
        row_idx += 1

    sub_cells = [
        ws1.cell(row=row_idx, column=1, value=cat + ' 小計'),
        ws1.cell(row=row_idx, column=2, value='(' + str(len(cat_items)) + ' 項)'),
        ws1.cell(row=row_idx, column=3, value=''),
        ws1.cell(row=row_idx, column=4, value=''),
        ws1.cell(row=row_idx, column=5, value=cat_qty),
        ws1.cell(row=row_idx, column=6, value=cat_amt),
        ws1.cell(row=row_idx, column=7, value=''),
        ws1.cell(row=row_idx, column=8, value=''),
    ]
    for c in sub_cells:
        c.fill = subtotal_fill
        c.font = Font(bold=True)
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
    sub_cells[4].number_format = '#,##0'
    sub_cells[5].number_format = '#,##0'
    grand_qty += cat_qty
    grand_amt += cat_amt
    row_idx += 1

total_cells = [
    ws1.cell(row=row_idx, column=1, value='總計'),
    ws1.cell(row=row_idx, column=2, value='72 項'),
    ws1.cell(row=row_idx, column=3, value=''),
    ws1.cell(row=row_idx, column=4, value=''),
    ws1.cell(row=row_idx, column=5, value=grand_qty),
    ws1.cell(row=row_idx, column=6, value=grand_amt),
    ws1.cell(row=row_idx, column=7, value=''),
    ws1.cell(row=row_idx, column=8, value=''),
]
for c in total_cells:
    c.fill = title_fill
    c.font = Font(bold=True, color='FFFFFF', size=12)
    c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center')
total_cells[4].number_format = '#,##0'
total_cells[5].number_format = '#,##0'

widths = {1: 10, 2: 13, 3: 38, 4: 9, 5: 10, 6: 14, 7: 8, 8: 26}
for col, w in widths.items():
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.freeze_panes = 'A3'

ws2 = out_wb.create_sheet('庫點明細')
ws2.merge_cells('A1:G1')
t2 = ws2.cell(row=1, column=1, value='門市周邊綠色不補號商品 — 各庫點庫存明細')
t2.font = Font(bold=True, color='FFFFFF', size=14)
t2.fill = title_fill
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

headers2 = ['分類', '商品型號', '品名規格', '庫點', '單價', '庫存數', '金額']
for col_idx, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col_idx, value=h)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws2.row_dimensions[2].height = 24

row_idx = 3
for cat in cat_order:
    cat_items = [r for r in results if r['cat'] == cat]
    for r in cat_items:
        code = r['code']
        rows_inv = inv.get(code, [])
        for ir in rows_inv:
            cells = [
                ws2.cell(row=row_idx, column=1, value=cat),
                ws2.cell(row=row_idx, column=2, value=code),
                ws2.cell(row=row_idx, column=3, value=ir['name'].replace('_x000D_', '').strip()),
                ws2.cell(row=row_idx, column=4, value=ir['site']),
                ws2.cell(row=row_idx, column=5, value=ir['price']),
                ws2.cell(row=row_idx, column=6, value=ir['qty']),
                ws2.cell(row=row_idx, column=7, value=ir['amt']),
            ]
            for c in cells:
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cells[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cells[4].number_format = '#,##0'
            cells[5].number_format = '#,##0'
            cells[6].number_format = '#,##0'
            if (ir['qty'] or 0) <= 0:
                cells[5].fill = zero_fill
            row_idx += 1

widths2 = {1: 10, 2: 13, 3: 36, 4: 10, 5: 10, 6: 10, 7: 12}
for col, w in widths2.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'A3'

out_path = '資料/綠色不補號商品_價格庫存彙總.xlsx'
out_wb.save(out_path)
print('已儲存:', out_path)
print('檔案大小:', os.path.getsize(out_path), 'bytes')
print('')
print('總庫存數量:', grand_qty)
print('總庫存金額: NT$', grand_amt)

old = '資料/綠色不補號商品清單_待補價格庫存.xlsx'
if os.path.exists(old):
    os.remove(old)
    print('已刪除舊暫存檔:', old)
