import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('資料/庫存明細20260415.xlsx', data_only=True)
ws = wb.active

# Size columns are col 8 (scheme indicator) through col 25 (approximately)
# col 9-25 are the actual size quantity columns

# Read header rows to map scheme -> columns with actual size labels
# Row 4: 0|F|... (scheme 0 uses col 9 'F')
# Row 5: 1|S|M|L|XL|2L|3L|4L (scheme 1 uses col 9-15)
# Row 6: 6|5|5.5|6|6.5|7|7.5|8.5 (scheme 6 uses col 9-15)

# For this analysis, I'll just count distinct size columns with non-zero data per product
# Size columns: 9 through ~24

SIZE_COL_START = 9
SIZE_COL_END = 25

# Green products
green_codes = [
    'A0102-00','A0103-40','A0107-20','A0107-30','A0108-33','A0111-70','A0112-50','A0112-90',
    'A0201-20','A0203-73','A0208-40','A0217-32','A0220-04','A0220-07','A0220-27','A0220-32',
    'A0220-69','A0222-67','A0223-07','A0223-40','A0223-70','A0224-40','A0224-70','A0225-50',
    'A0227-21','A0227-81','A0228-30','A0228-90','A0229-21','A0229-73','A0229-81',
    'A0231-46','A0231-50','A0232-72','A0232-90',
    'A0701-33','A0701-34','A0701-90',
    'L1208-90','L1209-79','L1209-90','L1215-50','L1215-59','L1215-70','L1215-90',
    'L1216-07','L1216-30','L1216-56','L1216-70','L1217-07','L1217-70',
    'L1218-20','L1218-30','L1218-55','L1219-55','L1219-90','L1223-33',
    'B0102','B0107','B0108',
    'A0620','A0626',
    'A2302-90','C1301','C1302','C1303',
    'A0002-69','A0002-90','A0003-40','A0003-73','A0003-90','A0003-99',
]
green_set = set(green_codes)

# For each green product, collect all size column data across all 庫點 rows
prod_size_data = {}  # code -> {scheme: xxx, cols_with_data: set, size_breakdown: {col_idx: total_qty}}
for r in range(7, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    if not code:
        continue
    code = str(code).strip()
    if code not in green_set:
        continue
    scheme = ws.cell(row=r, column=8).value
    scheme = str(scheme) if scheme is not None else ''

    if code not in prod_size_data:
        prod_size_data[code] = {'scheme': scheme, 'size_totals': {}}

    for c in range(SIZE_COL_START, SIZE_COL_END + 1):
        v = ws.cell(row=r, column=c).value
        if v is None or v == 0:
            continue
        try:
            v = int(v)
        except:
            continue
        if v != 0:
            prod_size_data[code]['size_totals'][c] = prod_size_data[code]['size_totals'].get(c, 0) + v

# Read size labels from header rows 4, 5, 6 for each scheme
# col 8 has the scheme indicator in each header row
size_labels_by_scheme = {}  # scheme -> {col: label}
for hr in [4, 5, 6]:
    scheme = str(ws.cell(row=hr, column=8).value)
    labels = {}
    for c in range(SIZE_COL_START, SIZE_COL_END + 1):
        v = ws.cell(row=hr, column=c).value
        if v is not None:
            labels[c] = str(v)
    size_labels_by_scheme[scheme] = labels

print('Size labels by scheme (from headers):')
for s, labels in size_labels_by_scheme.items():
    print(f'  scheme={s!r}: {labels}')

print('\n\nGreen products size analysis:')
print(f'{"商品型號":12s}{"方案":4s}{"尺寸欄數":8s} 分布')
print('-'*120)

results = []
for code in green_codes:
    if code not in prod_size_data:
        results.append({'code': code, 'scheme': '-', 'n_cols': 0, 'is_onesize': None})
        continue
    data = prod_size_data[code]
    scheme = data['scheme']
    cols_with_data = list(data['size_totals'].keys())
    n_cols = len(cols_with_data)

    # Show labels for these cols
    labels = size_labels_by_scheme.get(scheme, {})
    breakdown = []
    for c in sorted(cols_with_data):
        lbl = labels.get(c, f'col{c}')
        breakdown.append(f'{lbl}={data["size_totals"][c]}')

    # is_onesize if only 1 column has data OR scheme is '0' (F=free)
    is_onesize = (n_cols <= 1) or (scheme == '0')

    print(f'{code:12s}{scheme:4s}{n_cols:8d} {", ".join(breakdown)}')
    results.append({'code': code, 'scheme': scheme, 'n_cols': n_cols,
                    'is_onesize': is_onesize, 'breakdown': breakdown})

# Save
with open('/tmp/size_analysis.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)

print(f'\n\n不分尺寸商品 (單一尺寸欄 或 scheme=0):')
onesize = [r for r in results if r.get('is_onesize')]
for r in onesize:
    print(f'  {r["code"]}  方案={r["scheme"]}')

print(f'\n有尺寸商品 (多個尺寸欄):')
sized = [r for r in results if not r.get('is_onesize')]
for r in sized:
    print(f'  {r["code"]}  方案={r["scheme"]}  分布={r.get("breakdown")}')

print(f'\n總結: 不分尺寸={len(onesize)}, 有尺寸={len(sized)}')
