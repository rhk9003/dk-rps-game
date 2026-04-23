import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load existing aggregated results
with open('/tmp/green_results.json', 'r', encoding='utf-8') as f:
    results = json.load(f)

# Load size analysis
with open('/tmp/size_analysis.json', 'r', encoding='utf-8') as f:
    size_results = json.load(f)
size_lookup = {r['code']: r for r in size_results}

# Load inventory details
wb_inv = openpyxl.load_workbook('資料/庫存明細20260415.xlsx', data_only=True)
ws_inv = wb_inv.active

# Size column labels by scheme from header rows 4/5/6
# Only '0', '1', '6' are in headers, other schemes (A/B/C/D/E/G/Q) need inferring
# For scheme '0': col9='F'
# For scheme '1' (clothing): col9='S', col10='M', col11='L', col12='XL', col13='2L', col14='3L', col15='4L', col17='5L', col18='XS', col19='6XL'
# For scheme '6' (shoe US): col9='5', col10='5.5', col11='6', col12='6.5', col13='7', col14='7.5', col15='8.5', col17='9', col18='4.5'
# For other schemes — fallback to col index
SIZE_LABELS = {
    '0': {9: 'F'},
    '1': {9: 'S', 10: 'M', 11: 'L', 12: 'XL', 13: '2L', 14: '3L', 15: '4L', 17: '5L', 18: 'XS', 19: '6XL'},
    '6': {9: '5', 10: '5.5', 11: '6', 12: '6.5', 13: '7', 14: '7.5', 15: '8.5', 17: '9', 18: '4.5'},
}

inv = {}
for r in range(7, ws_inv.max_row + 1):
    code = ws_inv.cell(row=r, column=1).value
    if not code:
        continue
    code = str(code).strip()
    inv.setdefault(code, []).append({
        'name': str(ws_inv.cell(row=r, column=2).value or '').strip(),
        'site': str(ws_inv.cell(row=r, column=3).value or '').strip(),
        'price': ws_inv.cell(row=r, column=4).value,
        'qty': ws_inv.cell(row=r, column=6).value or 0,
        'amt': ws_inv.cell(row=r, column=7).value or 0,
    })

# Build workbook
out_wb = openpyxl.Workbook()
ws1 = out_wb.active
ws1.title = '綠色不補號彙總'

title_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
cat_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
subtotal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
zero_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
# 不分尺寸 fill (light blue)
onesize_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
# 有尺寸 fill (light orange)
multisize_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

ws1.merge_cells('A1:J1')
t = ws1.cell(row=1, column=1, value='門市周邊綠色不補號商品 — 價格、庫存、尺寸類型  (製表日 2026-04-15)')
t.font = Font(bold=True, color='FFFFFF', size=14)
t.fill = title_fill
t.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

headers = ['分類', '商品型號', '品名規格', '尺寸類型', '尺寸分布', '單價', '總庫存', '總庫存金額', '庫點數', '備註']
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col_idx, value=h)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws1.row_dimensions[2].height = 24

cat_order = ['襪子', '拖鞋', '射出拖鞋', '包包', '香氛', '鞋保養品', '鞋墊', '護具']
row_idx = 3
grand_qty = 0
grand_amt = 0
cnt_onesize = 0
cnt_multi = 0

for cat in cat_order:
    cat_items = [r for r in results if r['cat'] == cat]
    cat_qty = 0
    cat_amt = 0
    for r in cat_items:
        code = r['code']
        name = (r['inv_name'] if r['inv_name'] else r['pdf_name']).replace('_x000D_', '').strip()
        price = r['price']
        qty = r['total_qty']
        amt = r['total_amt']
        sites = r['n_sites']

        # Size info
        s = size_lookup.get(code, {})
        n_cols = s.get('n_cols', 0)
        scheme = s.get('scheme', '')

        # Re-compute breakdown with labels
        prod_rows = inv.get(code, [])
        size_totals = {}
        for pr in prod_rows:
            pass  # we already have breakdown list in s

        # Use breakdown from analysis
        breakdown_raw = s.get('breakdown', [])
        # Convert colN -> label when possible
        labels_for_scheme = SIZE_LABELS.get(scheme, {})
        breakdown_labeled = []
        for b in breakdown_raw:
            # format like "col12=227" or "F=2"
            if '=' in b:
                lbl, val = b.split('=', 1)
                if lbl.startswith('col'):
                    try:
                        ci = int(lbl[3:])
                        new_lbl = labels_for_scheme.get(ci, lbl)
                        breakdown_labeled.append(f'{new_lbl}={val}')
                    except:
                        breakdown_labeled.append(b)
                else:
                    breakdown_labeled.append(b)
            else:
                breakdown_labeled.append(b)
        breakdown_str = ', '.join(breakdown_labeled)

        is_onesize = n_cols <= 1
        size_type = '不分尺寸' if is_onesize else f'有尺寸 ({n_cols}種)'

        if is_onesize:
            cnt_onesize += 1
        else:
            cnt_multi += 1

        if qty > 0:
            note = ''
        else:
            note = f'庫存={qty} (退貨/盤虧)'

        cells = [
            ws1.cell(row=row_idx, column=1, value=cat),
            ws1.cell(row=row_idx, column=2, value=code),
            ws1.cell(row=row_idx, column=3, value=name),
            ws1.cell(row=row_idx, column=4, value=size_type),
            ws1.cell(row=row_idx, column=5, value=breakdown_str),
            ws1.cell(row=row_idx, column=6, value=price),
            ws1.cell(row=row_idx, column=7, value=qty),
            ws1.cell(row=row_idx, column=8, value=amt),
            ws1.cell(row=row_idx, column=9, value=sites),
            ws1.cell(row=row_idx, column=10, value=note),
        ]
        for c in cells:
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cells[0].fill = cat_fill
        cells[0].font = Font(bold=True)
        cells[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[4].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[9].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[5].number_format = '#,##0'
        cells[6].number_format = '#,##0'
        cells[7].number_format = '#,##0'

        # Color the 尺寸類型 cell based on type
        if is_onesize:
            cells[3].fill = onesize_fill
            cells[3].font = Font(bold=True, color='1F4E79')
        else:
            cells[3].fill = multisize_fill
            cells[3].font = Font(bold=True, color='7F6000')

        if qty <= 0:
            cells[5].fill = zero_fill

        cat_qty += qty
        cat_amt += amt
        row_idx += 1

    # Subtotal
    sub_cells = [
        ws1.cell(row=row_idx, column=1, value=cat + ' 小計'),
        ws1.cell(row=row_idx, column=2, value='(' + str(len(cat_items)) + ' 項)'),
    ] + [ws1.cell(row=row_idx, column=c, value='') for c in range(3, 7)] + [
        ws1.cell(row=row_idx, column=7, value=cat_qty),
        ws1.cell(row=row_idx, column=8, value=cat_amt),
        ws1.cell(row=row_idx, column=9, value=''),
        ws1.cell(row=row_idx, column=10, value=''),
    ]
    for c in sub_cells:
        c.fill = subtotal_fill
        c.font = Font(bold=True)
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
    sub_cells[6].number_format = '#,##0'
    sub_cells[7].number_format = '#,##0'
    grand_qty += cat_qty
    grand_amt += cat_amt
    row_idx += 1

# Grand total
total_cells = [
    ws1.cell(row=row_idx, column=1, value='總計'),
    ws1.cell(row=row_idx, column=2, value='72 項'),
    ws1.cell(row=row_idx, column=3, value=''),
    ws1.cell(row=row_idx, column=4, value=f'不分尺寸 {cnt_onesize} / 有尺寸 {cnt_multi}'),
    ws1.cell(row=row_idx, column=5, value=''),
    ws1.cell(row=row_idx, column=6, value=''),
    ws1.cell(row=row_idx, column=7, value=grand_qty),
    ws1.cell(row=row_idx, column=8, value=grand_amt),
    ws1.cell(row=row_idx, column=9, value=''),
    ws1.cell(row=row_idx, column=10, value=''),
]
for c in total_cells:
    c.fill = title_fill
    c.font = Font(bold=True, color='FFFFFF', size=12)
    c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center')
total_cells[6].number_format = '#,##0'
total_cells[7].number_format = '#,##0'

# Legend below total
row_idx += 2
ws1.cell(row=row_idx, column=1, value='說明:').font = Font(bold=True)
ws1.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=10)
ws1.cell(row=row_idx, column=2, value='「不分尺寸」= 該商品在所有庫點僅有「單一尺寸欄位」有庫存資料 (含 Free size / 一體適用)。「有尺寸 (N種)」= 該商品在庫存表中實際有 N 種不同尺寸的庫存紀錄,例如拖鞋 EU 37–44、鞋墊 S/M/L、襪子 女/男款等。')
ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws1.row_dimensions[row_idx].height = 30

widths = {1: 10, 2: 13, 3: 32, 4: 14, 5: 28, 6: 9, 7: 10, 8: 13, 9: 7, 10: 20}
for col, w in widths.items():
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.freeze_panes = 'A3'

# Sheet 2: 庫點明細 (unchanged but add size type column)
ws2 = out_wb.create_sheet('庫點明細')
ws2.merge_cells('A1:H1')
t2 = ws2.cell(row=1, column=1, value='門市周邊綠色不補號商品 — 各庫點庫存明細')
t2.font = Font(bold=True, color='FFFFFF', size=14)
t2.fill = title_fill
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

headers2 = ['分類', '商品型號', '品名規格', '尺寸類型', '庫點', '單價', '庫存數', '金額']
for col_idx, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col_idx, value=h)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws2.row_dimensions[2].height = 24

row_idx = 3
for cat in cat_order:
    cat_items = [r for r in results if r['cat'] == cat]
    for r in cat_items:
        code = r['code']
        s = size_lookup.get(code, {})
        n_cols = s.get('n_cols', 0)
        is_onesize = n_cols <= 1
        size_type = '不分尺寸' if is_onesize else f'有尺寸 ({n_cols}種)'
        rows_inv = inv.get(code, [])
        for ir in rows_inv:
            cells = [
                ws2.cell(row=row_idx, column=1, value=cat),
                ws2.cell(row=row_idx, column=2, value=code),
                ws2.cell(row=row_idx, column=3, value=ir['name'].replace('_x000D_', '').strip()),
                ws2.cell(row=row_idx, column=4, value=size_type),
                ws2.cell(row=row_idx, column=5, value=ir['site']),
                ws2.cell(row=row_idx, column=6, value=ir['price']),
                ws2.cell(row=row_idx, column=7, value=ir['qty']),
                ws2.cell(row=row_idx, column=8, value=ir['amt']),
            ]
            for c in cells:
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cells[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cells[5].number_format = '#,##0'
            cells[6].number_format = '#,##0'
            cells[7].number_format = '#,##0'
            if is_onesize:
                cells[3].fill = onesize_fill
            else:
                cells[3].fill = multisize_fill
            if (ir['qty'] or 0) <= 0:
                cells[6].fill = zero_fill
            row_idx += 1

widths2 = {1: 10, 2: 13, 3: 32, 4: 14, 5: 10, 6: 10, 7: 10, 8: 12}
for col, w in widths2.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'A3'

# Sheet 3: 不分尺寸商品清單
ws3 = out_wb.create_sheet('不分尺寸商品')
ws3.merge_cells('A1:G1')
t3 = ws3.cell(row=1, column=1, value='不分尺寸商品 ({} 項) — Free size / 一體適用'.format(cnt_onesize))
t3.font = Font(bold=True, color='FFFFFF', size=14)
t3.fill = title_fill
t3.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

headers3 = ['分類', '商品型號', '品名規格', '單價', '總庫存', '總庫存金額', '庫點數']
for col_idx, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=col_idx, value=h)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

row_idx = 3
for cat in cat_order:
    cat_items = [r for r in results if r['cat'] == cat]
    for r in cat_items:
        code = r['code']
        s = size_lookup.get(code, {})
        if s.get('n_cols', 0) > 1:
            continue
        name = (r['inv_name'] if r['inv_name'] else r['pdf_name']).replace('_x000D_', '').strip()
        cells = [
            ws3.cell(row=row_idx, column=1, value=cat),
            ws3.cell(row=row_idx, column=2, value=code),
            ws3.cell(row=row_idx, column=3, value=name),
            ws3.cell(row=row_idx, column=4, value=r['price']),
            ws3.cell(row=row_idx, column=5, value=r['total_qty']),
            ws3.cell(row=row_idx, column=6, value=r['total_amt']),
            ws3.cell(row=row_idx, column=7, value=r['n_sites']),
        ]
        for c in cells:
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cells[0].fill = cat_fill
        cells[0].font = Font(bold=True)
        cells[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[3].number_format = '#,##0'
        cells[4].number_format = '#,##0'
        cells[5].number_format = '#,##0'
        row_idx += 1

widths3 = {1: 10, 2: 13, 3: 40, 4: 9, 5: 10, 6: 13, 7: 8}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = 'A3'

# Sheet 4: 有尺寸商品清單
ws4 = out_wb.create_sheet('有尺寸商品')
ws4.merge_cells('A1:H1')
t4 = ws4.cell(row=1, column=1, value='有尺寸商品 ({} 項) — 多種尺寸分布'.format(cnt_multi))
t4.font = Font(bold=True, color='FFFFFF', size=14)
t4.fill = title_fill
t4.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 30

headers4 = ['分類', '商品型號', '品名規格', '尺寸分布', '單價', '總庫存', '總庫存金額', '庫點數']
for col_idx, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=col_idx, value=h)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

row_idx = 3
for cat in cat_order:
    cat_items = [r for r in results if r['cat'] == cat]
    for r in cat_items:
        code = r['code']
        s = size_lookup.get(code, {})
        if s.get('n_cols', 0) <= 1:
            continue
        name = (r['inv_name'] if r['inv_name'] else r['pdf_name']).replace('_x000D_', '').strip()
        scheme = s.get('scheme', '')
        breakdown_raw = s.get('breakdown', [])
        labels_for_scheme = SIZE_LABELS.get(scheme, {})
        breakdown_labeled = []
        for b in breakdown_raw:
            if '=' in b:
                lbl, val = b.split('=', 1)
                if lbl.startswith('col'):
                    try:
                        ci = int(lbl[3:])
                        new_lbl = labels_for_scheme.get(ci, lbl)
                        breakdown_labeled.append(f'{new_lbl}={val}')
                    except:
                        breakdown_labeled.append(b)
                else:
                    breakdown_labeled.append(b)
            else:
                breakdown_labeled.append(b)
        breakdown_str = ', '.join(breakdown_labeled)

        cells = [
            ws4.cell(row=row_idx, column=1, value=cat),
            ws4.cell(row=row_idx, column=2, value=code),
            ws4.cell(row=row_idx, column=3, value=name),
            ws4.cell(row=row_idx, column=4, value=breakdown_str),
            ws4.cell(row=row_idx, column=5, value=r['price']),
            ws4.cell(row=row_idx, column=6, value=r['total_qty']),
            ws4.cell(row=row_idx, column=7, value=r['total_amt']),
            ws4.cell(row=row_idx, column=8, value=r['n_sites']),
        ]
        for c in cells:
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cells[0].fill = cat_fill
        cells[0].font = Font(bold=True)
        cells[2].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[3].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cells[4].number_format = '#,##0'
        cells[5].number_format = '#,##0'
        cells[6].number_format = '#,##0'
        row_idx += 1

widths4 = {1: 10, 2: 13, 3: 32, 4: 30, 5: 9, 6: 10, 7: 13, 8: 8}
for col, w in widths4.items():
    ws4.column_dimensions[get_column_letter(col)].width = w
ws4.freeze_panes = 'A3'

out_path = '資料/綠色不補號商品_價格庫存彙總.xlsx'
out_wb.save(out_path)
print('已儲存:', out_path)
print('檔案大小:', os.path.getsize(out_path), 'bytes')
print('不分尺寸:', cnt_onesize, '項')
print('有尺寸:', cnt_multi, '項')
